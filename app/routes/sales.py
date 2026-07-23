from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_file
from ..database import get_db
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

bp = Blueprint("sales", __name__, url_prefix="/kassa")


@bp.route("/")
def checkout():
    return render_template("sales/checkout.html")


@bp.route("/afrekenen", methods=["POST"])
def record():
    """Process a completed sale from the checkout form."""
    payment_method = request.form.get("payment_method", "cash")
    note = request.form.get("note", "").strip() or None

    # Parse cart lines from form: line_product_id_N, line_name_N, line_qty_N, line_price_N
    lines = []
    i = 0
    while True:
        product_id = request.form.get(f"line_product_id_{i}")
        name = request.form.get(f"line_name_{i}")
        qty = request.form.get(f"line_qty_{i}")
        price = request.form.get(f"line_price_{i}")
        if name is None:
            break
        lines.append({
            "product_id": int(product_id) if product_id else None,
            "name": name,
            "qty": int(qty),
            "price": float(price),
            "subtotal": int(qty) * float(price),
        })
        i += 1

    if not lines:
        return redirect(url_for("sales.checkout"))

    total = sum(l["subtotal"] for l in lines)

    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO sales (total, payment_method, note) VALUES (?, ?, ?)",
            (total, payment_method, note),
        )
        sale_id = cur.lastrowid

        for l in lines:
            conn.execute(
                """INSERT INTO sale_lines
                     (sale_id, product_id, product_name, quantity, unit_price, subtotal)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (sale_id, l["product_id"], l["name"], l["qty"], l["price"], l["subtotal"]),
            )
            if l["product_id"]:
                conn.execute(
                    "UPDATE products SET stock = MAX(0, stock - ?), updated_at = datetime('now') WHERE id = ?",
                    (l["qty"], l["product_id"]),
                )

    return redirect(url_for("sales.confirmation", sale_id=sale_id))


@bp.route("/bevestiging/<int:sale_id>")
def confirmation(sale_id):
    with get_db() as conn:
        sale = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
        lines = conn.execute(
            "SELECT * FROM sale_lines WHERE sale_id = ?", (sale_id,)
        ).fetchall()
    return render_template("sales/confirmation.html", sale=sale, lines=lines)


@bp.route("/geschiedenis")
def history():
    from_date = request.args.get("van", date.today().isoformat())
    to_date = request.args.get("tot", date.today().isoformat())
    with get_db() as conn:
        sales = conn.execute(
            """SELECT * FROM sales
               WHERE date(created_at) BETWEEN ? AND ?
               ORDER BY created_at DESC""",
            (from_date, to_date),
        ).fetchall()
        day_total = sum(s["total"] for s in sales)
    return render_template(
        "sales/history.html",
        sales=sales,
        day_total=day_total,
        from_date=from_date,
        to_date=to_date,
    )


@bp.route("/<int:sale_id>/regels")
def sale_lines(sale_id):
    """HTMX endpoint: load sale detail lines."""
    with get_db() as conn:
        sale = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
        lines = conn.execute(
            "SELECT * FROM sale_lines WHERE sale_id = ?", (sale_id,)
        ).fetchall()
    return render_template("sales/lines_partial.html", sale=sale, lines=lines)


@bp.route("/export/xlsx")
def export_xlsx():
    """Export sales to Excel file."""
    from_date = request.args.get("van", date.today().isoformat())
    to_date = request.args.get("tot", date.today().isoformat())
    
    with get_db() as conn:
        sales = conn.execute(
            """SELECT * FROM sales
               WHERE date(created_at) BETWEEN ? AND ?
               ORDER BY created_at DESC""",
            (from_date, to_date),
        ).fetchall()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Verkopen"
    
    # Headers
    headers = ["ID", "Datum & tijd", "Betaalmethode", "Totaal"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="8B4789", end_color="8B4789", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Payment method labels
    labels = {'cash': 'Contant', 'payconiq': 'Payconiq', 'mixed': 'Gemengd'}
    
    # Add data
    for sale in sales:
        ws.append([
            sale['id'],
            sale['created_at'][:16],
            labels.get(sale['payment_method'], sale['payment_method']),
            sale['total'],
        ])
    
    # Format columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Right-align totals
    for row in ws.iter_rows(min_row=2, max_row=len(sales) + 1, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '€ #,##0.00'
    
    # Add totals row
    if sales:
        day_total = sum(s['total'] for s in sales)
        ws.append(["", "", f"Totaal ({len(sales)} verkopen)", day_total])
        
        # Style total row
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)
            if cell.column == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '€ #,##0.00'
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"verkopen_{from_date}_tot_{to_date}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
